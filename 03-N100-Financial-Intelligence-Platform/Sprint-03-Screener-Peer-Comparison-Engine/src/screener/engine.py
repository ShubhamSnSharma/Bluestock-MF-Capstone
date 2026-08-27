"""
Financial Screener & Composite Scoring Engine.
Sprint 3 — Screener & Peer Comparison Engine
"""

import os
import sqlite3
import yaml
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ScreenerEngine:
    def __init__(
        self,
        db_path: Optional[str] = None,
        config_path: Optional[str] = None,
        output_dir: Optional[str] = None
    ):
        sprint3_root = Path(__file__).resolve().parent.parent.parent
        project_root = sprint3_root.parent

        if db_path is None:
            self.db_path = project_root / "Sprint-01-Data-Foundation" / "nifty100.db"
        else:
            self.db_path = Path(db_path)

        if config_path is None:
            self.config_path = sprint3_root / "config" / "screener_config.yaml"
        else:
            self.config_path = Path(config_path)

        if output_dir is None:
            self.output_dir = sprint3_root / "output"
        else:
            self.output_dir = Path(output_dir)

        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.config = self.load_config()

    def load_config(self) -> Dict[str, Any]:
        """Loads analyst-editable screener configuration YAML."""
        if not self.config_path.exists():
            raise FileNotFoundError(f"Screener config not found at: {self.config_path}")
        with open(self.config_path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f)

    def load_universe_data(self, year: str = "2024") -> pd.DataFrame:
        """Loads and joins financial statement, valuation, and ratio metrics for the given year."""
        conn = sqlite3.connect(self.db_path)

        query = """
        SELECT
            fr.company_id,
            c.company_name,
            s.broad_sector,
            fr.year,
            fr.return_on_equity_pct AS roe,
            fr.return_on_capital_employed_pct AS roce,
            fr.net_profit_margin_pct AS npm,
            fr.operating_profit_margin_pct AS opm,
            fr.debt_to_equity AS de,
            fr.interest_coverage AS icr,
            fr.icr_label,
            fr.asset_turnover,
            fr.free_cash_flow_cr AS fcf,
            fr.cfo_quality_score AS cfo_pat_ratio,
            fr.revenue_cagr_3yr,
            fr.revenue_cagr_5yr,
            fr.pat_cagr_5yr,
            fr.eps_cagr_5yr,
            fr.dividend_payout_ratio_pct AS dividend_payout,
            pnl.sales,
            pnl.net_profit,
            mc.market_cap_crore AS market_cap,
            mc.pe_ratio AS pe,
            mc.pb_ratio AS pb,
            mc.dividend_yield_pct AS dividend_yield
        FROM financial_ratios fr
        JOIN companies c ON fr.company_id = c.id
        LEFT JOIN sectors s ON fr.company_id = s.company_id
        LEFT JOIN profitandloss pnl ON fr.company_id = pnl.company_id AND fr.year = pnl.year
        LEFT JOIN market_cap mc ON fr.company_id = mc.company_id AND fr.year = mc.year
        WHERE fr.year = ?
        """

        df = pd.read_sql_query(query, conn, params=[year])

        # Load previous year D/E for Turnaround Watch (YoY decline check)
        prev_year = str(int(year) - 1) if year.isdigit() else "2023"
        de_prev = pd.read_sql_query(
            "SELECT company_id, debt_to_equity AS de_prev FROM financial_ratios WHERE year = ?",
            conn, params=[prev_year]
        )
        df = pd.merge(df, de_prev, on="company_id", how="left")

        # Load 5-year ago FCF for FCF CAGR computation
        fcf_5yr_ago = pd.read_sql_query(
            "SELECT company_id, free_cash_flow_cr AS fcf_5yr_ago FROM financial_ratios WHERE year = ?",
            conn, params=[str(int(year) - 5) if year.isdigit() else "2019"]
        )
        df = pd.merge(df, fcf_5yr_ago, on="company_id", how="left")

        conn.close()

        # Compute Composite Quality Scores (Global and Sector-Relative)
        df = self.calculate_composite_scores(df)

        return df

    def winsorise_and_scale(self, series: pd.Series, lower_pct: float = 10.0, upper_pct: float = 90.0, invert: bool = False) -> pd.Series:
        """Winsorises a series at P10/P90 and scales linearly to 0-100."""
        valid_vals = series.dropna()
        if len(valid_vals) == 0:
            return pd.Series(50.0, index=series.index)

        p_low = np.percentile(valid_vals, lower_pct)
        p_high = np.percentile(valid_vals, upper_pct)

        # Clip extreme values
        clipped = series.clip(lower=p_low, upper=p_high)

        # Handle constant series
        if p_high == p_low:
            scaled = pd.Series(50.0, index=series.index)
        else:
            scaled = (clipped - p_low) / (p_high - p_low) * 100.0

        if invert:
            scaled = 100.0 - scaled

        # Fill missing values with neutral median (50.0)
        return scaled.fillna(50.0)

    def calculate_composite_scores(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Computes 0-100 Composite Quality Score:
        - 35% Profitability: ROE (15%), ROCE (10%), NPM (10%)
        - 30% Cash Quality: FCF CAGR (15%), CFO/PAT (10%), FCF Positive (5%)
        - 20% Growth: Revenue CAGR (10%), PAT CAGR (10%)
        - 15% Leverage: D/E (10%, inverted), ICR (5%)
        Also computes Sector-Relative Composite Score normalised within broad_sector.
        """
        df = df.copy()

        # Compute FCF 5-yr CAGR proxy / Score
        fcf_pos_flag = (df["fcf"] > 0).astype(float) * 100.0

        # Calculate raw FCF growth score (positive and expanding)
        fcf_growth_score = []
        for _, r in df.iterrows():
            f_now = r.get("fcf")
            f_old = r.get("fcf_5yr_ago")
            if pd.notnull(f_now) and pd.notnull(f_old) and f_old > 0 and f_now > 0:
                cagr = (((f_now / f_old) ** 0.2) - 1.0) * 100.0
                fcf_growth_score.append(cagr)
            elif pd.notnull(f_now) and f_now > 0:
                fcf_growth_score.append(10.0)
            else:
                fcf_growth_score.append(-10.0)
        df["fcf_cagr_score_raw"] = fcf_growth_score

        # Treat Debt Free / None ICR as high ICR for scoring
        icr_score_raw = df["icr"].copy()
        icr_score_raw[df["icr_label"] == "Debt Free"] = 200.0
        icr_score_raw[(df["de"] == 0) & (icr_score_raw.isnull())] = 200.0

        # Helper to compute composite score on a given dataframe subset
        def compute_subset_score(sub_df: pd.DataFrame) -> pd.Series:
            s_roe = self.winsorise_and_scale(sub_df["roe"])
            s_roce = self.winsorise_and_scale(sub_df["roce"])
            s_npm = self.winsorise_and_scale(sub_df["npm"])

            s_fcf_cagr = self.winsorise_and_scale(sub_df["fcf_cagr_score_raw"])
            s_cfo_pat = self.winsorise_and_scale(sub_df["cfo_pat_ratio"])
            s_fcf_flag = sub_df["fcf_pos_flag"]

            s_rev_cagr = self.winsorise_and_scale(sub_df["revenue_cagr_5yr"])
            s_pat_cagr = self.winsorise_and_scale(sub_df["pat_cagr_5yr"])

            s_de = self.winsorise_and_scale(sub_df["de"], invert=True)
            s_icr = self.winsorise_and_scale(sub_df["icr_score_raw"])

            profitability = 0.15 * s_roe + 0.10 * s_roce + 0.10 * s_npm
            cash_quality = 0.15 * s_fcf_cagr + 0.10 * s_cfo_pat + 0.05 * s_fcf_flag
            growth = 0.10 * s_rev_cagr + 0.10 * s_pat_cagr
            leverage = 0.10 * s_de + 0.05 * s_icr

            total_score = profitability + cash_quality + growth + leverage
            return total_score.round(2)

        df["fcf_pos_flag"] = fcf_pos_flag
        df["icr_score_raw"] = icr_score_raw

        # Global Composite Quality Score
        df["composite_quality_score"] = compute_subset_score(df)

        # Sector-Relative Composite Score
        df["sector_relative_score"] = df.groupby("broad_sector", group_keys=False).apply(
            lambda g: compute_subset_score(g), include_groups=False
        )

        return df

    def apply_filters(self, df: pd.DataFrame, filters: Dict[str, Any]) -> pd.DataFrame:
        """Applies customizable filter thresholds across supported metrics."""
        filtered = df.copy()

        # 1. ROE minimum
        if "roe_min" in filters and filters["roe_min"] is not None:
            filtered = filtered[filtered["roe"] >= filters["roe_min"]]

        # 2. D/E maximum (Automatically bypassed for Financials)
        if "de_max" in filters and filters["de_max"] is not None:
            max_de = filters["de_max"]
            filtered = filtered[(filtered["broad_sector"] == "Financials") | (filtered["de"] <= max_de)]

        # 3. FCF minimum
        if "fcf_min" in filters and filters["fcf_min"] is not None:
            filtered = filtered[filtered["fcf"] >= filters["fcf_min"]]

        # 4. Revenue CAGR 5yr minimum
        if "revenue_cagr_5yr_min" in filters and filters["revenue_cagr_5yr_min"] is not None:
            filtered = filtered[filtered["revenue_cagr_5yr"] >= filters["revenue_cagr_5yr_min"]]

        # 5. PAT CAGR 5yr minimum
        if "pat_cagr_5yr_min" in filters and filters["pat_cagr_5yr_min"] is not None:
            filtered = filtered[filtered["pat_cagr_5yr"] >= filters["pat_cagr_5yr_min"]]

        # 6. OPM minimum
        if "opm_min" in filters and filters["opm_min"] is not None:
            filtered = filtered[filtered["opm"] >= filters["opm_min"]]

        # 7. P/E maximum
        if "pe_max" in filters and filters["pe_max"] is not None:
            filtered = filtered[filtered["pe"] <= filters["pe_max"]]

        # 8. P/B maximum
        if "pb_max" in filters and filters["pb_max"] is not None:
            filtered = filtered[filtered["pb"] <= filters["pb_max"]]

        # 9. Dividend Yield minimum
        if "dividend_yield_min" in filters and filters["dividend_yield_min"] is not None:
            filtered = filtered[filtered["dividend_yield"] >= filters["dividend_yield_min"]]

        # 10. ICR minimum (Debt Free passes any ICR threshold)
        if "icr_min" in filters and filters["icr_min"] is not None:
            min_icr = filters["icr_min"]
            filtered = filtered[
                (filtered["icr_label"] == "Debt Free") |
                (filtered["de"] == 0) |
                (filtered["icr"] >= min_icr)
            ]

        # 11. Market Cap minimum
        if "market_cap_min" in filters and filters["market_cap_min"] is not None:
            filtered = filtered[filtered["market_cap"] >= filters["market_cap_min"]]

        # 12. Net Profit minimum
        if "net_profit_min" in filters and filters["net_profit_min"] is not None:
            filtered = filtered[filtered["net_profit"] >= filters["net_profit_min"]]

        # 13. EPS CAGR minimum
        if "eps_cagr_5yr_min" in filters and filters["eps_cagr_5yr_min"] is not None:
            filtered = filtered[filtered["eps_cagr_5yr"] >= filters["eps_cagr_5yr_min"]]

        # 14. Asset Turnover minimum
        if "asset_turnover_min" in filters and filters["asset_turnover_min"] is not None:
            filtered = filtered[filtered["asset_turnover"] >= filters["asset_turnover_min"]]

        # 15. Sales minimum
        if "sales_min" in filters and filters["sales_min"] is not None:
            filtered = filtered[filtered["sales"] >= filters["sales_min"]]

        # Special Preset Filter: Dividend Payout Maximum
        if "dividend_payout_max" in filters and filters["dividend_payout_max"] is not None:
            filtered = filtered[filtered["dividend_payout"] <= filters["dividend_payout_max"]]

        # Special Preset Filter: Revenue CAGR 3yr minimum
        if "revenue_cagr_3yr_min" in filters and filters["revenue_cagr_3yr_min"] is not None:
            filtered = filtered[filtered["revenue_cagr_3yr"] >= filters["revenue_cagr_3yr_min"]]

        # Special Preset Filter: D/E declining YoY
        if filters.get("de_declining_yoy") is True:
            filtered = filtered[
                (filtered["de"].notnull()) &
                (filtered["de_prev"].notnull()) &
                (filtered["de"] < filtered["de_prev"])
            ]

        return filtered.sort_values(by="composite_quality_score", ascending=False)

    def run_all_presets(self, df: Optional[pd.DataFrame] = None) -> Dict[str, pd.DataFrame]:
        """Runs all 6 preset screeners defined in configuration."""
        if df is None:
            df = self.load_universe_data()

        results = {}
        presets_config = self.config.get("presets", {})

        for preset_key, preset_data in presets_config.items():
            preset_name = preset_data.get("name", preset_key)
            filters = preset_data.get("filters", {})
            filtered_df = self.apply_filters(df, filters)
            results[preset_key] = filtered_df

        return results

    def export_screener_workbook(self, results: Dict[str, pd.DataFrame], output_file: Optional[str] = None):
        """
        Exports screener results to an Excel workbook with 6 sheets.
        Applies professional styling: green fill for passing cells, frozen headers, column auto-widths.
        """
        if output_file is None:
            out_path = self.output_dir / "screener_output.xlsx"
        else:
            out_path = Path(output_file)

        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        kpi_columns = [
            "company_id", "company_name", "broad_sector",
            "composite_quality_score", "sector_relative_score",
            "roe", "roce", "npm", "opm", "de", "icr",
            "fcf", "cfo_pat_ratio", "revenue_cagr_5yr", "pat_cagr_5yr", "eps_cagr_5yr",
            "pe", "pb", "dividend_yield", "market_cap"
        ]

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        presets_config = self.config.get("presets", {})

        for preset_key, preset_data in presets_config.items():
            preset_name = preset_data.get("name", preset_key)
            sheet_title = preset_name[:31] # Excel max sheet title length
            ws = wb.create_sheet(title=sheet_title)

            df_sheet = results.get(preset_key, pd.DataFrame())

            # Write header row
            ws.append(kpi_columns)
            for col_num in range(1, len(kpi_columns) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Write data rows
            if not df_sheet.empty:
                for row_idx, row in df_sheet.iterrows():
                    row_vals = []
                    for col in kpi_columns:
                        val = row.get(col)
                        if pd.isnull(val):
                            row_vals.append("N/A")
                        elif isinstance(val, (float, np.floating)):
                            row_vals.append(round(float(val), 2))
                        else:
                            row_vals.append(val)
                    ws.append(row_vals)

                    # Style cells
                    cur_row = ws.max_row
                    for col_num in range(1, len(kpi_columns) + 1):
                        cell = ws.cell(row=cur_row, column=col_num)
                        cell.border = thin_border
                        # Highlight KPI metrics with green fill
                        if col_num in (4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20):
                            cell.fill = pass_fill

            # Freeze top header row
            ws.freeze_panes = "A2"

            # Auto-fit column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(out_path)
        return out_path


if __name__ == "__main__":
    engine = ScreenerEngine()
    df_universe = engine.load_universe_data()
    print(f"Loaded universe: {len(df_universe)} companies.")
    results = engine.run_all_presets(df_universe)
    print("\nPreset Results Counts:")
    for preset, res_df in results.items():
        print(f"  {preset:<22}: {len(res_df):2d} companies")
    out_file = engine.export_screener_workbook(results)
    print(f"\nScreener output saved to: {out_file}")
