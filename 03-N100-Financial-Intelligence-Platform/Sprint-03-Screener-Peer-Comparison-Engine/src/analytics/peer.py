"""
Peer Percentile Ranking, Comparison Report, and Radar Visualization Engine.
Sprint 3 — Screener & Peer Comparison Engine
"""

import os
import sqlite3
import math
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class PeerComparisonEngine:
    def __init__(
        self,
        db_path: Optional[str] = None,
        output_dir: Optional[str] = None,
        reports_dir: Optional[str] = None
    ):
        sprint3_root = Path(__file__).resolve().parent.parent.parent
        project_root = sprint3_root.parent

        if db_path is None:
            self.db_path = project_root / "Sprint-01-Data-Foundation" / "nifty100.db"
        else:
            self.db_path = Path(db_path)

        if output_dir is None:
            self.output_dir = sprint3_root / "output"
        else:
            self.output_dir = Path(output_dir)

        if reports_dir is None:
            self.reports_dir = sprint3_root / "reports" / "radar_charts"
        else:
            self.reports_dir = Path(reports_dir)

        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.reports_dir.mkdir(parents=True, exist_ok=True)

        self.ranking_metrics = [
            ("roe", "return_on_equity_pct", False),
            ("roce", "return_on_capital_employed_pct", False),
            ("npm", "net_profit_margin_pct", False),
            ("de", "debt_to_equity", True),  # Inverted: lower D/E = higher percentile
            ("fcf", "free_cash_flow_cr", False),
            ("pat_cagr_5yr", "pat_cagr_5yr", False),
            ("revenue_cagr_5yr", "revenue_cagr_5yr", False),
            ("eps_cagr_5yr", "eps_cagr_5yr", False),
            ("icr", "interest_coverage", False),
            ("asset_turnover", "asset_turnover", False),
        ]

    def ensure_peer_percentiles_table(self, conn: sqlite3.Connection):
        """Creates peer_percentiles table and indexes if not exists."""
        cursor = conn.cursor()
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS peer_percentiles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id TEXT NOT NULL,
            peer_group_name TEXT NOT NULL,
            metric TEXT NOT NULL,
            value REAL,
            percentile_rank REAL,
            year TEXT NOT NULL,
            FOREIGN KEY (company_id) REFERENCES companies(id) ON DELETE CASCADE,
            UNIQUE (company_id, peer_group_name, metric, year)
        );
        """)
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_pp_company ON peer_percentiles(company_id);")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_pp_group ON peer_percentiles(peer_group_name);")
        conn.commit()

    def load_peer_data(self, year: str = "2024") -> Tuple[pd.DataFrame, pd.DataFrame]:
        """Loads peer groups mapping and financial metrics for all constituents."""
        conn = sqlite3.connect(self.db_path)

        pg_df = pd.read_sql_query("SELECT * FROM peer_groups", conn)

        query = """
        SELECT
            fr.company_id,
            c.company_name,
            s.broad_sector,
            fr.year,
            fr.return_on_equity_pct,
            fr.return_on_capital_employed_pct,
            fr.net_profit_margin_pct,
            fr.operating_profit_margin_pct,
            fr.debt_to_equity,
            fr.interest_coverage,
            fr.icr_label,
            fr.asset_turnover,
            fr.free_cash_flow_cr,
            fr.cfo_quality_score,
            fr.capex_intensity_pct,
            fr.revenue_cagr_5yr,
            fr.pat_cagr_5yr,
            fr.eps_cagr_5yr,
            fr.dividend_payout_ratio_pct,
            fr.composite_quality_score,
            pnl.sales,
            pnl.net_profit,
            mc.market_cap_crore,
            mc.pe_ratio,
            mc.pb_ratio,
            mc.dividend_yield_pct
        FROM financial_ratios fr
        JOIN companies c ON fr.company_id = c.id
        LEFT JOIN sectors s ON fr.company_id = s.company_id
        LEFT JOIN profitandloss pnl ON fr.company_id = pnl.company_id AND fr.year = pnl.year
        LEFT JOIN market_cap mc ON fr.company_id = mc.company_id AND fr.year = mc.year
        WHERE fr.year = ?
        """
        all_metrics_df = pd.read_sql_query(query, conn, params=[year])
        conn.close()

        return pg_df, all_metrics_df

    def compute_peer_percentiles(self, pg_df: pd.DataFrame, metrics_df: pd.DataFrame, year: str = "2024") -> pd.DataFrame:
        """
        Computes percentile ranks for 10 metrics across all 11 peer groups.
        Populates SQLite peer_percentiles table.
        """
        conn = sqlite3.connect(self.db_path)
        conn.execute("PRAGMA foreign_keys = ON;")
        self.ensure_peer_percentiles_table(conn)

        # Merge peer groups with metrics
        peer_merged = pd.merge(pg_df, metrics_df, on="company_id", how="inner")

        percentile_records = []

        # Calculate percentiles within each peer group
        for group_name, group_data in peer_merged.groupby("peer_group_name"):
            n_companies = len(group_data)

            for metric_code, col_name, invert in self.ranking_metrics:
                # Prepare metric series
                series = group_data[col_name].copy()

                # Treat Debt Free ICR as high value
                if metric_code == "icr":
                    series[group_data["icr_label"] == "Debt Free"] = 999.0
                    series[(group_data["debt_to_equity"] == 0) & (series.isnull())] = 999.0

                # Compute percentile rank (0 to 100%)
                if invert:
                    # Invert D/E: lower D/E = higher percentile
                    # Rank negative of series so smallest value gets rank 1
                    ranks = (-series).rank(ascending=True, pct=True, method="average") * 100.0
                else:
                    ranks = series.rank(ascending=True, pct=True, method="average") * 100.0

                for idx, row in group_data.iterrows():
                    cid = row["company_id"]
                    val = row.get(col_name)
                    rank_val = ranks.loc[idx]

                    if pd.isnull(val):
                        final_rank = None
                        final_val = None
                    else:
                        final_val = float(val)
                        final_rank = round(float(rank_val), 2)

                    percentile_records.append({
                        "company_id": cid,
                        "peer_group_name": group_name,
                        "metric": metric_code,
                        "value": final_val,
                        "percentile_rank": final_rank,
                        "year": year
                    })

        pp_df = pd.DataFrame(percentile_records)

        # Populate SQLite peer_percentiles table
        cursor = conn.cursor()
        cursor.execute("DELETE FROM peer_percentiles WHERE year = ?", (year,))
        for _, r in pp_df.iterrows():
            cursor.execute("""
            INSERT INTO peer_percentiles (company_id, peer_group_name, metric, value, percentile_rank, year)
            VALUES (?, ?, ?, ?, ?, ?)
            """, (r["company_id"], r["peer_group_name"], r["metric"], r["value"], r["percentile_rank"], r["year"]))

        conn.commit()
        conn.close()

        return pp_df

    def generate_radar_charts(self, pg_df: pd.DataFrame, metrics_df: pd.DataFrame):
        """
        Generates 8-axis polar radar charts for all 92 companies:
        - For peer-group companies: Company polygon vs Peer Group average dashed outline.
        - For standalone companies: Company polygon vs Nifty 100 universe average.
        Saves PNGs to reports/radar_charts/{company_id}_radar.png.
        """
        radar_axes = [
            ("ROE", "return_on_equity_pct", False),
            ("ROCE", "return_on_capital_employed_pct", False),
            ("NPM", "net_profit_margin_pct", False),
            ("D/E Health", "debt_to_equity", True),
            ("FCF Score", "free_cash_flow_cr", False),
            ("PAT CAGR 5Y", "pat_cagr_5yr", False),
            ("Rev CAGR 5Y", "revenue_cagr_5yr", False),
            ("Quality Score", "composite_quality_score", False),
        ]
        axis_labels = [a[0] for a in radar_axes]
        n_axes = len(radar_axes)
        angles = np.linspace(0, 2 * np.pi, n_axes, endpoint=False).tolist()
        angles += angles[:1] # Close the circle

        # Map peer group per company
        comp_to_group = pg_df.set_index("company_id")["peer_group_name"].to_dict()
        peer_merged = pd.merge(pg_df, metrics_df, on="company_id", how="inner")

        # Precompute min/max bounds for 0-100 scaling on radar axes
        bounds = {}
        for label, col, invert in radar_axes:
            vals = metrics_df[col].dropna()
            p10 = np.percentile(vals, 10.0) if len(vals) > 0 else 0.0
            p90 = np.percentile(vals, 90.0) if len(vals) > 0 else 100.0
            bounds[col] = (p10, p90, invert)

        def scale_val(val: Optional[float], col: str) -> float:
            if pd.isnull(val):
                return 50.0
            p10, p90, invert = bounds[col]
            if p90 == p10:
                s = 50.0
            else:
                clipped = max(p10, min(float(val), p90))
                s = (clipped - p10) / (p90 - p10) * 100.0
            return 100.0 - s if invert else s

        # 1. Compute peer group average scaled profiles
        group_profiles = {}
        for group_name, group_data in peer_merged.groupby("peer_group_name"):
            avg_profile = []
            for label, col, invert in radar_axes:
                scaled_col = [scale_val(v, col) for v in group_data[col]]
                avg_profile.append(np.mean(scaled_col))
            avg_profile += avg_profile[:1]
            group_profiles[group_name] = avg_profile

        # 2. Compute Nifty 100 universe average profile for standalone companies
        universe_profile = []
        for label, col, invert in radar_axes:
            scaled_col = [scale_val(v, col) for v in metrics_df[col]]
            universe_profile.append(np.mean(scaled_col))
        universe_profile += universe_profile[:1]

        # 3. Generate chart for each company
        chart_count = 0
        for _, row in metrics_df.iterrows():
            cid = row["company_id"]
            cname = row.get("company_name", cid)
            group_name = comp_to_group.get(cid)

            # Company profile
            comp_profile = [scale_val(row.get(col), col) for label, col, invert in radar_axes]
            comp_profile += comp_profile[:1]

            # Create plot
            fig, ax = plt.subplots(figsize=(7, 7), subplot_kw=dict(polar=True))
            ax.set_theta_offset(np.pi / 2)
            ax.set_theta_direction(-1)

            # Draw axis lines and labels
            plt.xticks(angles[:-1], axis_labels, color="#2C3E50", size=10, weight="bold")
            ax.set_rlabel_position(0)
            plt.yticks([25, 50, 75, 100], ["25", "50", "75", "100"], color="#7F8C8D", size=8)
            plt.ylim(0, 100)

            # Plot company polygon
            ax.plot(angles, comp_profile, color="#1F77B4", linewidth=2.2, linestyle="solid", label=f"{cid} ({cname[:20]})")
            ax.fill(angles, comp_profile, color="#1F77B4", alpha=0.35)

            # Plot reference benchmark overlay
            if group_name:
                ref_profile = group_profiles[group_name]
                ref_label = f"Peer Avg: {group_name}"
                ax.plot(angles, ref_profile, color="#E67E22", linewidth=2.0, linestyle="dashed", label=ref_label)
            else:
                ref_profile = universe_profile
                ref_label = "Nifty 100 Universe Avg"
                ax.plot(angles, ref_profile, color="#95A5A6", linewidth=1.8, linestyle="dashed", label=ref_label)

            plt.title(f"{cid} — Financial Performance Radar\n({group_name or 'Standalone Constituent'})", size=13, weight="bold", pad=20, color="#1F4E78")
            plt.legend(loc="upper right", bbox_to_anchor=(1.25, 1.1), fontsize=9)
            plt.tight_layout()

            chart_path = self.reports_dir / f"{cid}_radar.png"
            plt.savefig(chart_path, dpi=150, bbox_inches="tight")
            plt.close(fig)
            chart_count += 1

        return chart_count

    def export_peer_comparison_workbook(self, pg_df: pd.DataFrame, metrics_df: pd.DataFrame, pp_df: pd.DataFrame, output_file: Optional[str] = None):
        """
        Generates output/peer_comparison.xlsx with exactly 11 sheets (one per peer group).
        Features:
        - 20 metric columns + percentile ranks.
        - Color-coded percentile ranks: >=75th green, 25th-75th yellow, <=25th red.
        - Gold/Amber row highlight for benchmark company (is_benchmark = 1).
        - Peer group median summary row at bottom.
        """
        if output_file is None:
            out_path = self.output_dir / "peer_comparison.xlsx"
        else:
            out_path = Path(output_file)

        wb = openpyxl.Workbook()
        wb.remove(wb.active) # Remove default sheet

        # 20 Metric Columns
        display_metrics = [
            ("ROE (%)", "return_on_equity_pct", "roe"),
            ("ROCE (%)", "return_on_capital_employed_pct", "roce"),
            ("NPM (%)", "net_profit_margin_pct", "npm"),
            ("OPM (%)", "operating_profit_margin_pct", None),
            ("D/E", "debt_to_equity", "de"),
            ("ICR", "interest_coverage", "icr"),
            ("Asset Turnover", "asset_turnover", "asset_turnover"),
            ("FCF (₹ Cr)", "free_cash_flow_cr", "fcf"),
            ("CFO/PAT", "cfo_quality_score", None),
            ("CapEx Intensity (%)", "capex_intensity_pct", None),
            ("Rev CAGR 5Y (%)", "revenue_cagr_5yr", "revenue_cagr_5yr"),
            ("PAT CAGR 5Y (%)", "pat_cagr_5yr", "pat_cagr_5yr"),
            ("EPS CAGR 5Y (%)", "eps_cagr_5yr", "eps_cagr_5yr"),
            ("P/E Ratio", "pe_ratio", None),
            ("P/B Ratio", "pb_ratio", None),
            ("Dividend Yield (%)", "dividend_yield_pct", None),
            ("Dividend Payout (%)", "dividend_payout_ratio_pct", None),
            ("Sales (₹ Cr)", "sales", None),
            ("Net Profit (₹ Cr)", "net_profit", None),
            ("Composite Score", "composite_quality_score", None),
        ]

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        benchmark_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Gold/Amber
        median_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Light Blue
        median_font = Font(name="Calibri", size=11, bold=True, color="1F4E78")

        # Percentile rank conditional colors
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # >= 75%
        green_font = Font(color="006100", bold=True)
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # 25-75%
        yellow_font = Font(color="9C6500")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # <= 25%
        red_font = Font(color="9C0006")

        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        peer_merged = pd.merge(pg_df, metrics_df, on="company_id", how="inner")

        # Index percentiles by (company_id, metric)
        pp_dict = pp_df.set_index(["company_id", "metric"])["percentile_rank"].to_dict()

        for group_name, group_data in peer_merged.groupby("peer_group_name"):
            sheet_title = group_name[:31]
            ws = wb.create_sheet(title=sheet_title)

            # Build headers: Company Info + 20 Metrics (with percentile column for 10 ranking metrics)
            header_row = ["Company ID", "Company Name", "Is Benchmark"]
            for m_label, m_col, m_code in display_metrics:
                header_row.append(m_label)
                if m_code:
                    header_row.append(f"{m_label} Pctile")

            ws.append(header_row)
            for col_idx in range(1, len(header_row) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Data rows
            row_start = 2
            for _, r in group_data.iterrows():
                cid = r["company_id"]
                is_bm = bool(r.get("is_benchmark", 0))

                row_vals = [cid, r.get("company_name", cid), "YES (Benchmark)" if is_bm else "NO"]

                for m_label, m_col, m_code in display_metrics:
                    val = r.get(m_col)
                    if pd.isnull(val):
                        row_vals.append("N/A")
                    elif isinstance(val, (float, np.floating)):
                        row_vals.append(round(float(val), 2))
                    else:
                        row_vals.append(val)

                    if m_code:
                        pctile_val = pp_dict.get((cid, m_code))
                        if pctile_val is not None and pd.notnull(pctile_val):
                            row_vals.append(f"{pctile_val:.1f}%")
                        else:
                            row_vals.append("N/A")

                ws.append(row_vals)
                cur_row = ws.max_row

                # Apply styling
                for col_idx in range(1, len(header_row) + 1):
                    cell = ws.cell(row=cur_row, column=col_idx)
                    cell.border = thin_border

                    # Highlight benchmark row
                    if is_bm:
                        cell.fill = benchmark_fill

                    # Format percentile rank cells
                    header_name = header_row[col_idx - 1]
                    if "Pctile" in header_name and cell.value and cell.value != "N/A":
                        try:
                            p_num = float(str(cell.value).replace("%", ""))
                            if p_num >= 75.0:
                                cell.fill = green_fill
                                cell.font = green_font
                            elif p_num <= 25.0:
                                cell.fill = red_fill
                                cell.font = red_font
                            else:
                                cell.fill = yellow_fill
                                cell.font = yellow_font
                        except Exception:
                            pass

            # Add Peer Group Median Summary Row
            median_row_vals = ["PEER MEDIAN", f"Median of {len(group_data)} Peers", "—"]
            for m_label, m_col, m_code in display_metrics:
                num_vals = group_data[m_col].dropna()
                if len(num_vals) > 0:
                    med_val = round(float(np.median(num_vals)), 2)
                    median_row_vals.append(med_val)
                else:
                    median_row_vals.append("N/A")

                if m_code:
                    median_row_vals.append("50.0% (Median)")

            ws.append(median_row_vals)
            med_row_idx = ws.max_row
            for col_idx in range(1, len(header_row) + 1):
                cell = ws.cell(row=med_row_idx, column=col_idx)
                cell.font = median_font
                cell.fill = median_fill
                cell.border = thin_border

            # Freeze headers
            ws.freeze_panes = "D2"

            # Auto column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

        wb.save(out_path)
        return out_path

    def run(self) -> Dict[str, Any]:
        """Runs the complete peer percentile ranking and report export workflow."""
        pg_df, metrics_df = self.load_peer_data()
        pp_df = self.compute_peer_percentiles(pg_df, metrics_df)
        chart_count = self.generate_radar_charts(pg_df, metrics_df)
        peer_wb_path = self.export_peer_comparison_workbook(pg_df, metrics_df, pp_df)

        return {
            "peer_groups_count": pg_df["peer_group_name"].nunique(),
            "peer_companies_count": pg_df["company_id"].nunique(),
            "percentiles_records_count": len(pp_df),
            "radar_charts_generated": chart_count,
            "peer_comparison_workbook": str(peer_wb_path)
        }


if __name__ == "__main__":
    engine = PeerComparisonEngine()
    summary = engine.run()
    print("Sprint 3 Peer Comparison Engine Execution Summary:")
    for k, v in summary.items():
        print(f"  {k}: {v}")
