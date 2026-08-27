"""
Unit tests for Screener Engine, 6 Presets, and Composite Quality Scores.
Sprint 3 — Screener & Peer Comparison Engine
"""

import pytest
import pandas as pd
import numpy as np
import openpyxl
from pathlib import Path
from src.screener.engine import ScreenerEngine


@pytest.fixture(scope="module")
def screener_engine():
    return ScreenerEngine()


@pytest.fixture(scope="module")
def universe_df(screener_engine):
    return screener_engine.load_universe_data(year="2024")


def test_screener_config_loaded(screener_engine):
    config = screener_engine.config
    assert "presets" in config
    assert "quality_compounder" in config["presets"]
    assert "value_pick" in config["presets"]
    assert "growth_accelerator" in config["presets"]
    assert "dividend_champion" in config["presets"]
    assert "debt_free_blue_chip" in config["presets"]
    assert "turnaround_watch" in config["presets"]


def test_universe_data_loaded(universe_df):
    assert len(universe_df) == 92
    assert "company_id" in universe_df.columns
    assert "roe" in universe_df.columns
    assert "de" in universe_df.columns
    assert "fcf" in universe_df.columns
    assert "composite_quality_score" in universe_df.columns
    assert "sector_relative_score" in universe_df.columns


def test_composite_score_bounds(universe_df):
    scores = universe_df["composite_quality_score"].dropna()
    assert len(scores) == 92
    assert (scores >= 0.0).all()
    assert (scores <= 100.0).all()


def test_sector_relative_score_bounds(universe_df):
    sec_scores = universe_df["sector_relative_score"].dropna()
    assert len(sec_scores) == 92
    assert (sec_scores >= 0.0).all()
    assert (sec_scores <= 100.0).all()


def test_winsorisation_and_scaling(screener_engine):
    # Test series with extreme outliers
    s = pd.Series([1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 1000.0])
    scaled = screener_engine.winsorise_and_scale(s, lower_pct=10.0, upper_pct=90.0)
    assert (scaled >= 0.0).all()
    assert (scaled <= 100.0).all()
    # Outlier 1000 should be capped at 100.0
    assert scaled.iloc[-1] == pytest.approx(100.0)


def test_de_financials_carve_out(screener_engine):
    sample_df = pd.DataFrame([
        {"company_id": "HDFCBANK", "broad_sector": "Financials", "de": 7.5, "roe": 16.0, "composite_quality_score": 80.0},
        {"company_id": "TATAMOTORS", "broad_sector": "Consumer Discretionary", "de": 7.5, "roe": 16.0, "composite_quality_score": 70.0},
        {"company_id": "INFY", "broad_sector": "Information Technology", "de": 0.1, "roe": 25.0, "composite_quality_score": 90.0}
    ])
    # Apply filter de_max = 1.0
    res = screener_engine.apply_filters(sample_df, {"de_max": 1.0})
    passed_ids = res["company_id"].tolist()
    # HDFCBANK should pass (bypassed because Financials), INFY passes (0.1 < 1.0), TATAMOTORS filtered out
    assert "HDFCBANK" in passed_ids
    assert "INFY" in passed_ids
    assert "TATAMOTORS" not in passed_ids


def test_debt_free_icr_screening(screener_engine):
    sample_df = pd.DataFrame([
        {"company_id": "COALINDIA", "broad_sector": "Energy", "de": 0.0, "icr": None, "icr_label": "Debt Free", "composite_quality_score": 85.0},
        {"company_id": "HIGH_ICR", "broad_sector": "Industrials", "de": 0.2, "icr": 25.0, "icr_label": None, "composite_quality_score": 75.0},
        {"company_id": "LOW_ICR", "broad_sector": "Materials", "de": 1.5, "icr": 1.2, "icr_label": None, "composite_quality_score": 50.0}
    ])
    # Apply icr_min = 5.0
    res = screener_engine.apply_filters(sample_df, {"icr_min": 5.0})
    passed_ids = res["company_id"].tolist()
    # COALINDIA (Debt Free) and HIGH_ICR pass, LOW_ICR filtered out
    assert "COALINDIA" in passed_ids
    assert "HIGH_ICR" in passed_ids
    assert "LOW_ICR" not in passed_ids


def test_all_six_presets_run(screener_engine, universe_df):
    results = screener_engine.run_all_presets(universe_df)
    assert len(results) == 6
    assert "quality_compounder" in results
    assert "value_pick" in results
    assert "growth_accelerator" in results
    assert "dividend_champion" in results
    assert "debt_free_blue_chip" in results
    assert "turnaround_watch" in results

    # Verify each preset returns non-empty result dataframe
    for preset_name, res_df in results.items():
        assert isinstance(res_df, pd.DataFrame)
        assert len(res_df) > 0, f"Preset {preset_name} returned 0 results"


def test_quality_compounder_criteria(screener_engine, universe_df):
    results = screener_engine.run_all_presets(universe_df)
    qc_df = results["quality_compounder"]
    # Check that all returned companies satisfy ROE >= 15, FCF >= 0, Rev CAGR 5yr >= 10
    assert (qc_df["roe"] >= 15.0).all()
    assert (qc_df["fcf"] >= 0.0).all()
    assert (qc_df["revenue_cagr_5yr"] >= 10.0).all()


def test_screener_output_excel_export(screener_engine, universe_df):
    results = screener_engine.run_all_presets(universe_df)
    out_path = screener_engine.export_screener_workbook(results)
    assert out_path.exists()

    wb = openpyxl.load_workbook(out_path)
    # Exactly 6 sheets
    assert len(wb.sheetnames) == 6
    expected_sheets = ["Quality Compounder", "Value Pick", "Growth Accelerator", "Dividend Champion", "Debt-Free Blue Chip", "Turnaround Watch"]
    for expected in expected_sheets:
        assert expected in wb.sheetnames
