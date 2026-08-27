"""
Unit tests for Peer Comparison Engine, Percentiles, and Visualization.
Sprint 3 — Screener & Peer Comparison Engine
"""

import pytest
import sqlite3
import pandas as pd
import openpyxl
from pathlib import Path
from src.analytics.peer import PeerComparisonEngine


@pytest.fixture(scope="module")
def peer_engine():
    return PeerComparisonEngine()


def test_peer_groups_loaded(peer_engine):
    pg_df, metrics_df = peer_engine.load_peer_data()
    assert pg_df["peer_group_name"].nunique() == 11
    assert pg_df["company_id"].nunique() == 56
    assert len(metrics_df) == 92


def test_peer_percentiles_table_population(peer_engine):
    summary = peer_engine.run()
    assert summary["peer_groups_count"] == 11
    assert summary["peer_companies_count"] == 56
    assert summary["percentiles_records_count"] == 560

    conn = sqlite3.connect(peer_engine.db_path)
    count = conn.execute("SELECT COUNT(*) FROM peer_percentiles").fetchone()[0]
    conn.close()
    assert count == 560


def test_inverse_de_percentile(peer_engine):
    # In peer_percentiles, lower D/E should have higher percentile rank
    conn = sqlite3.connect(peer_engine.db_path)
    de_ranks = pd.read_sql_query(
        "SELECT company_id, peer_group_name, value, percentile_rank FROM peer_percentiles WHERE metric = 'de'",
        conn
    )
    conn.close()

    for group_name, group_data in de_ranks.groupby("peer_group_name"):
        if len(group_data) > 1:
            # Company with min D/E should have max percentile rank
            min_de_comp = group_data.loc[group_data["value"].idxmin()]
            max_rank_comp = group_data.loc[group_data["percentile_rank"].idxmax()]
            assert min_de_comp["company_id"] == max_rank_comp["company_id"], f"Failed for group {group_name}"


def test_it_services_highest_roe_percentile(peer_engine):
    conn = sqlite3.connect(peer_engine.db_path)
    it_roe = pd.read_sql_query(
        "SELECT company_id, value, percentile_rank FROM peer_percentiles WHERE peer_group_name = 'IT Services' AND metric = 'roe'",
        conn
    )
    conn.close()

    assert len(it_roe) == 5
    # The company with highest ROE value should have the highest percentile rank (100.0%)
    max_val_comp = it_roe.loc[it_roe["value"].idxmax()]
    max_rank_comp = it_roe.loc[it_roe["percentile_rank"].idxmax()]
    assert max_val_comp["company_id"] == max_rank_comp["company_id"]
    assert max_rank_comp["percentile_rank"] == pytest.approx(100.0)


def test_missing_peer_group_handling(peer_engine):
    # 36 companies have no peer group; verify they do not cause errors
    pg_df, metrics_df = peer_engine.load_peer_data()
    comp_in_pg = set(pg_df["company_id"])
    non_pg_comps = metrics_df[~metrics_df["company_id"].isin(comp_in_pg)]
    assert len(non_pg_comps) == 36


def test_peer_comparison_excel_sheets(peer_engine):
    wb_path = peer_engine.output_dir / "peer_comparison.xlsx"
    assert wb_path.exists()

    wb = openpyxl.load_workbook(wb_path)
    assert len(wb.sheetnames) == 11

    expected_groups = [
        'Private Banks', 'Public Sector Banks', 'IT Services', 'Pharmaceuticals',
        'Automobiles', 'Life Insurance', 'Oil & Gas', 'Power & Utilities',
        'Steel', 'FMCG', 'Consumer Finance'
    ]
    for g in expected_groups:
        assert g in wb.sheetnames


def test_radar_charts_generated(peer_engine):
    chart_files = list(peer_engine.reports_dir.glob("*_radar.png"))
    assert len(chart_files) == 92
