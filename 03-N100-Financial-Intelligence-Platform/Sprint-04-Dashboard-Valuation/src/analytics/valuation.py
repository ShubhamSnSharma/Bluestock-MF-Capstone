"""
valuation.py — FCF Yield & Overvaluation Flag Engine
Sprint 4 — Day 26

Computes for all 92 companies (FY2024):
  • FCF Yield % = FCF / Market Cap × 100
  • Sector median P/E
  • PE flags: Caution (P/E > sector_median × 1.5)
              Discount (P/E < sector_median × 0.7)
              Fair (otherwise)

Generates:
  • output/valuation_summary.xlsx  — 92-row full report
  • output/valuation_flags.csv     — Caution + Discount companies only
"""

from __future__ import annotations

import logging
import sqlite3
from pathlib import Path

import pandas as pd

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
_SPRINT4_ROOT  = Path(__file__).resolve().parents[2]   # Sprint-04-Dashboard-Valuation/
_PLATFORM_ROOT = Path(__file__).resolve().parents[3]   # 03-N100-Financial-Intelligence-Platform/
DB_PATH        = (
    _PLATFORM_ROOT
    / "Sprint-01-Data-Foundation"
    / "nifty100.db"
)
OUTPUT_DIR     = _SPRINT4_ROOT / "output"

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
CAUTION_MULTIPLIER  = 1.5
DISCOUNT_MULTIPLIER = 0.7


# ---------------------------------------------------------------------------
# Data loaders (plain sqlite3 — no Streamlit cache needed here)
# ---------------------------------------------------------------------------

def _load_db() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load all required tables from nifty100.db."""
    if not DB_PATH.exists():
        raise FileNotFoundError(f"Database not found: {DB_PATH}")

    con = sqlite3.connect(f"file:{DB_PATH}?mode=ro", uri=True)
    try:
        companies = pd.read_sql_query(
            "SELECT id AS company_id, company_name FROM companies", con
        )
        sectors = pd.read_sql_query(
            "SELECT company_id, broad_sector, sub_sector, market_cap_category FROM sectors", con
        )
        market_cap = pd.read_sql_query(
            """
            SELECT company_id, year, market_cap_crore, enterprise_value_crore,
                   pe_ratio, pb_ratio, ev_ebitda, dividend_yield_pct
            FROM market_cap
            WHERE year = '2024'
            """,
            con,
        )
        ratios = pd.read_sql_query(
            """
            SELECT company_id, year,
                   free_cash_flow_cr,
                   return_on_equity_pct,
                   revenue_cagr_5yr,
                   pat_cagr_5yr,
                   composite_quality_score
            FROM financial_ratios
            WHERE year = '2024'
            """,
            con,
        )
    finally:
        con.close()

    return companies, sectors, market_cap, ratios


# ---------------------------------------------------------------------------
# Core valuation engine
# ---------------------------------------------------------------------------

def compute_valuation() -> pd.DataFrame:
    """
    Compute valuation metrics and flags for all 92 companies.

    Returns a DataFrame with columns:
        company_id, company_name, sector, P/E, P/B, EV/EBITDA,
        FCF_yield_pct, 5yr_median_PE, PE_vs_sector_median_pct, flag
    """
    companies, sectors, market_cap, ratios = _load_db()

    # ── Merge base ────────────────────────────────────────────────────────
    df = (
        companies
        .merge(sectors, on="company_id", how="left")
        .merge(market_cap, on="company_id", how="left")
        .merge(
            ratios[["company_id", "free_cash_flow_cr", "composite_quality_score"]],
            on="company_id",
            how="left",
        )
    )

    # ── FCF Yield ─────────────────────────────────────────────────────────
    def _fcf_yield(row: pd.Series) -> float | None:
        """FCF Yield % = (FCF / Market Cap) × 100"""
        fcf = row["free_cash_flow_cr"]
        mc  = row["market_cap_crore"]
        if pd.isna(fcf) or pd.isna(mc) or mc <= 0:
            return None
        return round((fcf / mc) * 100, 4)

    df["FCF_yield_pct"] = df.apply(_fcf_yield, axis=1)

    # ── Sector median P/E ─────────────────────────────────────────────────
    sector_pe_median = (
        df.groupby("broad_sector")["pe_ratio"]
        .median()
        .reset_index()
        .rename(columns={"pe_ratio": "sector_median_pe"})
    )
    df = df.merge(sector_pe_median, on="broad_sector", how="left")

    # Compute 5yr median PE — market_cap table only has 2019–2024 so we use available years
    con = sqlite3.connect(f"file:{DB_PATH}?mode=ro", uri=True)
    mc_all = pd.read_sql_query(
        "SELECT company_id, year, pe_ratio FROM market_cap WHERE year != 'TTM'", con
    )
    con.close()

    pe_5yr_median = (
        mc_all[mc_all["year"].astype(str) >= "2020"]
        .groupby("company_id")["pe_ratio"]
        .median()
        .reset_index()
        .rename(columns={"pe_ratio": "5yr_median_PE"})
    )
    df = df.merge(pe_5yr_median, on="company_id", how="left")

    # ── P/E vs sector median % ────────────────────────────────────────────
    def _pe_vs_median(row: pd.Series) -> float | None:
        pe  = row["pe_ratio"]
        med = row["sector_median_pe"]
        if pd.isna(pe) or pd.isna(med) or med == 0:
            return None
        return round(((pe - med) / med) * 100, 2)

    df["PE_vs_sector_median_pct"] = df.apply(_pe_vs_median, axis=1)

    # ── Valuation flag ────────────────────────────────────────────────────
    def _flag(row: pd.Series) -> str:
        pe  = row["pe_ratio"]
        med = row["sector_median_pe"]
        if pd.isna(pe) or pd.isna(med) or med == 0:
            return "Fair"                               # insufficient data → neutral
        if pe > med * CAUTION_MULTIPLIER:
            return "Caution"
        if pe < med * DISCOUNT_MULTIPLIER:
            return "Discount"
        return "Fair"

    df["flag"] = df.apply(_flag, axis=1)

    # ── Final column selection & rename ───────────────────────────────────
    result = df[[
        "company_id", "company_name", "broad_sector",
        "pe_ratio", "pb_ratio", "ev_ebitda",
        "FCF_yield_pct", "5yr_median_PE", "sector_median_pe",
        "PE_vs_sector_median_pct", "flag",
    ]].copy()

    result = result.rename(columns={
        "broad_sector"  : "sector",
        "pe_ratio"      : "P/E",
        "pb_ratio"      : "P/B",
        "ev_ebitda"     : "EV/EBITDA",
        "sector_median_pe": "sector_median_PE",
    })

    result = result.sort_values("company_id").reset_index(drop=True)

    flag_counts = result["flag"].value_counts().to_dict()
    logger.info(
        "Valuation computed: %d companies | flags: %s",
        len(result),
        flag_counts,
    )

    return result


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

def write_outputs(df: pd.DataFrame) -> tuple[Path, Path]:
    """
    Write valuation_summary.xlsx and valuation_flags.csv to the output/ directory.
    Returns (xlsx_path, csv_path).
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    xlsx_path = OUTPUT_DIR / "valuation_summary.xlsx"
    csv_path  = OUTPUT_DIR / "valuation_flags.csv"

    # ── valuation_summary.xlsx ────────────────────────────────────────────
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Valuation Summary")

        ws = writer.sheets["Valuation Summary"]

        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter

        # Header formatting
        header_fill   = PatternFill("solid", fgColor="1F2937")
        header_font   = Font(bold=True, color="E6EDF3")
        header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin          = Side(style="thin", color="374151")
        border        = Border(bottom=thin)

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = header_align
            cell.border    = border

        # Column widths
        col_widths = {
            "A": 12, "B": 32, "C": 22, "D": 8, "E": 8,
            "F": 10, "G": 14, "H": 14, "I": 16, "J": 20, "K": 10,
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Flag colouring
        FILL_CAUTION  = PatternFill("solid", fgColor="3D2F00")
        FILL_DISCOUNT = PatternFill("solid", fgColor="1A4731")
        FONT_CAUTION  = Font(color="E3B341", bold=True)
        FONT_DISCOUNT = Font(color="3FB950", bold=True)

        flag_col_idx = df.columns.tolist().index("flag") + 1  # 1-indexed

        for row in ws.iter_rows(min_row=2):
            flag_cell = row[flag_col_idx - 1]
            if flag_cell.value == "Caution":
                flag_cell.fill = FILL_CAUTION
                flag_cell.font = FONT_CAUTION
            elif flag_cell.value == "Discount":
                flag_cell.fill = FILL_DISCOUNT
                flag_cell.font = FONT_DISCOUNT

        ws.freeze_panes = "A2"

    logger.info("Written: %s", xlsx_path)

    # ── valuation_flags.csv ───────────────────────────────────────────────
    flags_df = df[df["flag"].isin(["Caution", "Discount"])].copy()
    flags_df.to_csv(csv_path, index=False)
    logger.info(
        "Written: %s (%d flagged companies)", csv_path, len(flags_df)
    )

    return xlsx_path, csv_path


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def run() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
    logger.info("Starting valuation engine …")

    df = compute_valuation()

    # Summary stats
    flag_counts = df["flag"].value_counts()
    logger.info("Flag distribution:\n%s", flag_counts.to_string())
    logger.info("FCF yield stats:\n%s", df["FCF_yield_pct"].describe().to_string())
    logger.info("Total companies: %d", len(df))

    xlsx, csv = write_outputs(df)

    print(f"\n✅ valuation_summary.xlsx → {xlsx}")
    print(f"✅ valuation_flags.csv   → {csv}")
    print(f"\nFlag distribution:\n{flag_counts.to_string()}")
    print(f"\nRows: {len(df)} | Flagged (Caution+Discount): {flag_counts.get('Caution',0)+flag_counts.get('Discount',0)}")


if __name__ == "__main__":
    run()
